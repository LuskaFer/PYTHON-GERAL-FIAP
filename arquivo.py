# 1. Ler as notas inseridas na planilha notas.xlsx
# 2. Efetuar os cálculos e emitir um relatório mostrando a nota mínima necessária na Global Solution para aprovação

# Média Semestral (40% - Média de Checkpoints (as duas notas mais altas dentre os CPs) + Challenges) + (60% - Global Solution)

# Média Anual = (40% - Média 1º Semestre) + (60% - Média do 2º Semestre)

from openpyxl import load_workbook

wb = load_workbook('notas.xlsx') #abrindo o Workbook .xlsx
ws = wb['Planilha1'] #selecionando a planilha 'x' dentro do Workbook
# for line in ws: # iterando em todas as linhas da Plan1
#     print(line[0])

# ----- Cálculo -----
nota_minima = 6

 #Aluna 1
a1_cp1 = float(ws['E4'].value)
a1_cp2 = float(ws['G4'].value)
a1_ch3 = float(ws['I4'].value)
a1_ch4 = float(ws['J4'].value)

a1_nota1 = (a1_cp1 + a1_cp2 + a1_ch3 + a1_ch4) / 4
print("nota 1 = ", a1_nota1)

# Cálculo da Global Solution -> 60% que faltam
a1_nota2 = (nota_minima - (a1_nota1 * 0.4)) / 0.6
print("nota 2 = ", a1_nota2)

# Média semestral
# ms = ((0.4 * a1_nota1) + (0.6 * a1_nota2))
# print('ms ', round(ms, 2))

#Média anual:
# sem1 = float(ws['C4'].value)
# mf = (0.4 * sem1) + (0.6 * sem2)

# Aluna 2
a2_cp1 = float(ws['E5'].value)
a2_cp2 = float(ws['G5'].value)
a2_ch3 = float(ws['I5'].value)
a2_ch4 = float(ws['J5'].value)

a2_nota1 = (a2_cp1 + a2_cp2 + a2_ch3 + a2_ch4) / 4
print(a2_nota1)

a2_nota2 = (nota_minima - (a2_nota1 * 0.4)) / 0.6
print("nota 2 = ", a2_nota2)

# Aluno 3
a3_cp1 = float(ws['E6'].value)
a3_cp2 = float(ws['G6'].value)
a3_ch3 = float(ws['I6'].value)
a3_ch4 = float(ws['J6'].value)

a3_nota1 = (a3_cp1 + a3_cp2 + a3_ch3 + a3_ch4) / 4
print(a3_nota1)

a3_nota2 = (nota_minima - (a3_nota1 * 0.4)) / 0.6
print("nota 2 = ", a3_nota2)

# Aluno 4
a4_cp1 = float(ws['E7'].value)
a4_cp2 = float(ws['G7'].value)
a4_ch3 = float(ws['I7'].value)
a4_ch4 = float(ws['J7'].value)

a4_nota1 = (a4_cp1 + a4_cp2 + a4_ch3 + a4_ch4) / 4
print(a4_nota1)

a4_nota2 = (nota_minima - (a4_nota1 * 0.4)) / 0.6
print("nota 2 = ", a4_nota2)

# Aluno 5
a5_cp1 = float(ws['E8'].value)
a5_cp2 = float(ws['G8'].value)
a5_ch3 = float(ws['I8'].value)
a5_ch4 = float(ws['J8'].value)

a5_nota1 = (a5_cp1 + a5_cp2 + a5_ch3 + a5_ch4) / 4
print(a5_nota1)

a5_nota2 = (nota_minima - (a5_nota1 * 0.4)) / 0.6
print("nota 2 = ", a5_nota2)

# ----- Relatório -----

# Aluna 1
print(ws['A4'].value, " ", ws['B4'].value) # RM e nome
print(ws['A1'].value, ws['C1'].value) # Disciplina
print("Semestre 1: ", ws['C4'].value) # Média do 1º semestre
print(" " * 15, "Semestre 2")
print("Checkpoints (média) :", (a1_cp1 + a1_cp2) / 2)
print("Challenge: ", (a1_ch3 + a1_ch4) / 2)
print("Nota mínima na Global Solution para aprovação: ", round(a1_nota2, 2))
print()

# Aluna 2
print(ws['A5'].value, " ", ws['B5'].value) # RM e nome
print(ws['A1'].value, ws['C1'].value) # Disciplina
print("Semestre 1: ", ws['C5'].value) # Média do 1º semestre
print(" " * 15, "Semestre 2")
print("Checkpoints (média) :", (a2_cp1 + a2_cp2) / 2)
print("Challenge: ", (a2_ch3 + a2_ch4) / 2)
print("Nota mínima na Global Solution para aprovação: ", round(a2_nota2, 2))
print()

# Aluno 3
print(ws['A6'].value, " ", ws['B6'].value) # RM e nome
print(ws['A1'].value, ws['C1'].value) # Disciplina
print("Semestre 1: ", ws['C6'].value) # Média do 1º semestre
print(" " * 15, "Semestre 2")
print("Checkpoints (média) :", (a3_cp1 + a3_cp2) / 2)
print("Challenge: ", (a3_ch3 + a3_ch4) / 2)
print("Nota mínima na Global Solution para aprovação: ", round(a3_nota2, 2))
print()

# Aluno 4
print(ws['A7'].value, " ", ws['B7'].value) # RM e nome
print(ws['A1'].value, ws['C1'].value) # Disciplina
print("Semestre 1: ", ws['C7'].value) # Média do 1º semestre
print(" " * 15, "Semestre 2")
print("Checkpoints (média) :", (a4_cp1 + a4_cp2) / 2)
print("Challenge: ", (a4_ch3 + a4_ch4) / 2)
print("Nota mínima na Global Solution para aprovação: ", round(a4_nota2, 2))
print()

# Aluno 5
print(ws['A8'].value, " ", ws['B8'].value) # RM e nome
print(ws['A1'].value, ws['C1'].value) # Disciplina
print("Semestre 1: ", ws['C8'].value) # Média do 1º semestre
print(" " * 15, "Semestre 2")
print("Checkpoints (média) :", (a5_cp1 + a5_cp2) / 2)
print("Challenge: ", (a5_ch3 + a5_ch4) / 2)
print("Nota mínima na Global Solution para aprovação: ", round(a5_nota2, 2))



